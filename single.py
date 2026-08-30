# -*- coding: utf-8 -*-
"""
报关自动化工具 v2.8（分表选择修复 / 性能增强 / 标准四舍五入 / PPH不参与）
============================================================================
本版相对 v2.6 的改动（计税公式一行未动）：

【核心修复】合同多分表时选错表：
    你的合同 xlsx 有 "attachment" 和 "CI" 两个分表，数据实际在 attachment。
    原代码默认读第 1 个分表(CI)，导致读不到 BM/PPN/PPH，全部算出 0。
    现在 read_rows() 会：
        1) 优先找名字含 "attachment" / "附件" 的分表（忽略大小写/空格/下划线）
        2) 找不到再找名字含 "attach" 的
        3) 都找不到才回退到原行为（第 1 个分表 / active）
    并在日志打印【分表列表 -> 选中】，方便核对。

【性能增强（解决合同多时"未响应"）】
    A. 合同文件只扫描 1 次建索引字典，匹配从 O(行数×文件数) 降为 O(行数)
    B. 日志批量刷新（满 20 行或 0.3 秒刷一次），不再每行都重绘 Tk 控件
    C. 损坏文件魔术字节预判（读前 8 字节），避免 xlrd/openpyxl 解析损坏包卡 5~10 秒
    D. 表头定位增强：优先找【同时含 AMOUNT 且 含 BM 且 含 PPN/PPH】的行

【计税公式（与 v2.6 完全一致，未动）】
    每行 BM  = AMOUNT × BM%
    每行 PPN = (AMOUNT + BM) × PPN%
    每行总税 = BM + PPN                ← 不含 PPH
    汇总 = 所有商品行求和 → round2 标准四舍五入（第3位5进位）
    增值税金额：代码不填，总表用 =ROUND((发票金额列 + 关税列) * 0.11, 2)
============================================================================
"""
import os
import re
import sys
import glob
import time
import traceback
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import load_workbook

try:
    import xlrd
except Exception:
    xlrd = None

HEADER_ROW = 2
DATA_START_ROW = 3
DEBUG = True

# ===== 日志批量刷新参数 =====
LOG_FLUSH_LINES = 20      # 攒够这么多行刷一次界面
LOG_FLUSH_SECONDS = 0.3   # 或超过这么多秒刷一次

_LOG_CACHE = []
_LAST_FLUSH = 0.0


# ==================== 工具 ====================
def round2(x):
    """标准四舍五入到 2 位小数（第3位5进位），解决 Python round 的银行家舍入"""
    if x is None:
        return 0.0
    try:
        return float(Decimal(str(x)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    except Exception:
        return round(float(x), 2)


def dbg(box, s, force=False):
    """日志：终端实时打印；界面批量刷新（force=True 立即刷）"""
    global _LAST_FLUSH
    if not DEBUG:
        return
    print(s, flush=True)
    if box is None:
        return
    _LOG_CACHE.append(s)
    now = time.time()
    if force or len(_LOG_CACHE) >= LOG_FLUSH_LINES or (now - _LAST_FLUSH) > LOG_FLUSH_SECONDS:
        _flush_log(box, now)


def _flush_log(box, now=None):
    global _LAST_FLUSH
    if not _LOG_CACHE:
        return
    try:
        box.insert("end", "\n".join(_LOG_CACHE) + "\n")
        _LOG_CACHE.clear()
        box.see("end")
        _LAST_FLUSH = now if now is not None else time.time()
    except Exception:
        pass


def cell_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v)


def norm(s):
    if s is None:
        return ""
    s = str(s).replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s)
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"[_\-]", "", s)
    return s.strip().upper()


def to_num(x):
    """转数字；'免/免征/none'→0；识别%符号"""
    if x is None or str(x).strip() == "":
        return 0.0, False
    s = str(x).replace(",", "").replace(" ", "").replace("％", "%").strip()
    low = s.lower()
    is_pct = "%" in s
    if any(k in low for k in ["免", "none", "na", "n/a"]):
        return 0.0, False
    try:
        v = float(s.replace("%", ""))
    except Exception:
        return 0.0, False
    return v, is_pct


def to_rate(x):
    """单元格 → 小数比率。值>1视为百分比(需/100)，否则已是小数"""
    val, is_pct = to_num(x)
    if val is None:
        return 0.0
    if is_pct or (val > 1 and val < 100):
        return val / 100.0
    return val


def looks_like_xls_or_xlsx(path):
    """
    魔术字节预判：xlsx=PK(zip头), xls=D0CF11E0(OLE2头)
    损坏/伪装文件读前 8 字节即判定，不进 xlrd/openpyxl 解析，避免卡 5~10 秒
    """
    try:
        with open(path, "rb") as f:
            head = f.read(8)
        if head[:2] == b"PK":
            return True
        if head[:4] == b"\xd0\xcf\x11\xe0":
            return True
        return False
    except Exception:
        return False


# ==================== 分表选择（本版核心修复）====================
def pick_sheet_name(sheetnames, box=None):
    """
    从分表列表中挑出应该读取的那个：
        1) 名字含 ATTACHMENT / 附件   ← 你的合同数据在 attachment
        2) 名字含 ATTACH
        3) 都没有 → 返回 None（调用方回退到 active / 第1个表，行为同 v2.6）
    """
    if not sheetnames:
        return None
    # 1) 强匹配：attachment / 附件
    for sn in sheetnames:
        n = norm(sn)
        if "ATTACHMENT" in n or "附件" in n:
            return sn
    # 2) 次匹配：attach
    for sn in sheetnames:
        n = norm(sn)
        if "ATTACH" in n:
            return sn
    return None


# ==================== 总表列定位 ====================
def header_score(row):
    s = " ".join(cell_text(v) for v in row).upper()
    keys = ["IV&PL", "INVOICE", "发票", "关税", "总税", "备注",
            "AMOUNT", "DESCRIPTION", "BM", "PPN", "PPH", "金额"]
    return sum(1 for k in keys if k.replace("&", "") in s or k in s)


def find_header_row(ws, max_check=10):
    best = (0, 1)
    for r in range(1, min(max_check, ws.max_row) + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        sc = header_score(row)
        if sc > best[0]:
            best = (sc, r)
    return best[1]


def col_by_names(ws, header_row, name_groups):
    headers = [norm(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    used_cols = set()
    for groups in name_groups:
        for name in groups:
            nn = norm(name)
            if not nn:
                continue
            for i, h in enumerate(headers, start=1):
                if i in used_cols:
                    continue
                if h and (nn in h or h in nn or nn.replace("&", "") in h):
                    used_cols.add(i)
                    return i
    return None


# ==================== 合同读取 ====================
def read_rows(path, box=None):
    """
    读取合同行数据。
    ★ 本版核心：优先读 attachment/附件 分表，读不到才回退原行为。
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".xls":
            if xlrd is None:
                raise RuntimeError("需要 xlrd 读取 .xls（pip install xlrd==1.2.0）")
            wb = xlrd.open_workbook(path)
            names = list(wb.sheet_names())
            chosen = pick_sheet_name(names, box)
            dbg(box, f"  分表列表={names} → 选中={chosen if chosen else (names[0] if names else '?')}")
            sh = wb.sheet_by_name(chosen) if chosen else wb.sheet_by_index(0)
            return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]

        # .xlsx / .xlsm
        wb = load_workbook(path, data_only=True, read_only=True)
        names = list(wb.sheetnames)
        chosen = pick_sheet_name(names, box)
        dbg(box, f"  分表列表={names} → 选中={chosen if chosen else (names[0] if names else '?')}")
        ws = wb[chosen] if chosen else wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    except Exception as e:
        # .xls 实为 xlsx 的兜底（保持 v2.6 行为）
        if ext == ".xls" and "not supported" in str(e).lower():
            try:
                wb = load_workbook(path, data_only=True, read_only=True)
                names = list(wb.sheetnames)
                chosen = pick_sheet_name(names, box)
                dbg(box, f"  [兜底]分表列表={names} → 选中={chosen if chosen else (names[0] if names else '?')}")
                ws = wb[chosen] if chosen else wb.active
                rows = [list(row) for row in ws.iter_rows(values_only=True)]
                wb.close()
                return rows
            except Exception:
                pass
        raise


def find_table_header(rows, max_check=25):
    """
    找税率表头行（本版增强，三级优先）：
        1) 同时含 AMOUNT 且 含 BM 且 含(PPN或PPH)   ← 最准确
        2) 含 BM 且 含(PPN或PPH)                    ← v2.6 原逻辑
        3) 含 AMOUNT                                ← 兜底
    """
    # 1) 最强
    for i, row in enumerate(rows[:max_check], start=1):
        texts = " ".join(cell_text(v).upper() for v in row)
        if "AMOUNT" in texts and "BM" in texts and ("PPN" in texts or "PPH" in texts):
            return i
    # 2) 次强（v2.6 原逻辑）
    for i, row in enumerate(rows[:max_check], start=1):
        texts = " ".join(cell_text(v).upper() for v in row)
        if "BM" in texts and ("PPN" in texts or "PPH" in texts):
            return i
    # 3) 含 AMOUNT
    for i, row in enumerate(rows[:max_check], start=1):
        texts = " ".join(cell_text(v).upper() for v in row)
        if "AMOUNT" in texts:
            return i
    # 4) 最后兜底
    last = 1
    for i, row in enumerate(rows[:max_check], start=1):
        texts = " ".join(cell_text(v).upper() for v in row)
        if any(k in texts for k in ["BM", "PPN", "PPH", "税额", "合计税率", "税率"]):
            last = i
    return last


def find_col_in_row(hdr, *names, used=None):
    for n in names:
        nn = norm(n)
        if not nn:
            continue
        for i, h in enumerate(hdr, start=1):
            if used and i in used:
                continue
            if h and (nn in h or h in nn or nn.replace("&", "") in h):
                return i
    return None


def get_grand_amount(rows, header_row, amt_col):
    """合计行(GRAND/TOTAL/合计)的 AMOUNT 值"""
    for r in range(header_row + 1, len(rows) + 1):
        row = rows[r - 1]
        texts = " ".join(cell_text(v).upper() for v in row)
        if any(k in texts for k in ["GRAND", "合计", "TOTAL", "小计"]):
            if amt_col:
                n, _ = to_num(row[amt_col - 1])
                if n and n > 0:
                    return n, r
            best = None
            for v in row:
                n, _ = to_num(v)
                if n and n > 100 and (best is None or n > best):
                    best = n
            if best:
                return best, r
    return None, None


def get_amount_fallback(rows):
    best = None
    for row in rows:
        for v in row:
            n, _ = to_num(v)
            if n and n > 100 and (best is None or n > best):
                best = n
    return best


def read_contract(path, box=None):
    """
    返回 {bm, ppn, pph, total_ex_pph, grand_amount}
    公式：
        每行 BM  = AMOUNT × BM%
        每行 PPN = (AMOUNT + BM) × PPN%
        每行总税 = BM + PPN            ← 不含 PPH
        汇总 = 所有商品行求和，再用 round2 标准四舍五入
    """
    dbg(box, f"\n  读取合同: {os.path.basename(path)}")

    # 损坏文件预检（避免解析损坏包卡死）
    if not looks_like_xls_or_xlsx(path):
        dbg(box, f"  ⚠ 跳过（非标准Excel格式，疑似损坏）: {os.path.basename(path)}", force=True)
        return None

    try:
        rows = read_rows(path, box)
    except Exception as e:
        dbg(box, f"  读取失败: {e}", force=True)
        return None
    dbg(box, f"  行数={len(rows)}")

    hr = 6
    try:
        hr = find_table_header(rows, max_check=25)
    except Exception:
        hr = 6
    if hr < 1 or hr > len(rows):
        hr = 1

    hdr = [norm(v) for v in rows[hr - 1]]
    dbg(box, f"  税率表头行=第{hr}行")
    dbg(box, f"  表头: {[cell_text(v) for v in rows[hr-1]]}")

    used = set()
    amt_col = find_col_in_row(hdr, "AMOUNT", "金额", "CIF", used=used)
    bm_col  = find_col_in_row(hdr, "BM可免", "BM", used=used) or find_col_in_row(hdr, "BM", used=used)
    ppn_col = find_col_in_row(hdr, "PPN", "VAT", "增值税", used=used)
    pph_col = find_col_in_row(hdr, "PPH", "WHT", used=used)   # 只读不用
    dbg(box, f"  列: AMOUNT={amt_col} BM={bm_col} PPN={ppn_col} PPH={pph_col}(不进表)")

    grand, grand_row = get_grand_amount(rows, hr, amt_col)
    if not grand:
        grand = get_amount_fallback(rows)
    dbg(box, f"  合计行=第{grand_row}行  GRAND_AMOUNT={grand}")

    # ===== 逐行：BM=H×M，总税行=H×M+(H+H×M)×O（不含 PPH）=====
    sum_bm = 0.0
    sum_tax = 0.0
    for r in range(hr + 1, len(rows) + 1):
        if r == grand_row:
            continue
        row = rows[r - 1]
        amt, _ = to_num(row[amt_col - 1]) if amt_col else (0.0, False)
        if not amt or amt <= 0:
            continue
        bm_raw = row[bm_col - 1] if bm_col else None
        if bm_raw is not None and str(bm_raw).strip() in ["免", "免征"]:
            bm_r = 0.0
        else:
            bm_r = to_rate(bm_raw)
        ppn_r = to_rate(row[ppn_col - 1]) if ppn_col else 0.0

        bm = amt * bm_r
        sum_bm += bm
        sum_tax += bm + (amt + bm) * ppn_r

    info = {
        "amount": grand,
        "bm": round2(sum_bm),
        "ppn": round2(sum_tax - sum_bm),
        "pph": 0.0,
        "total_ex_pph": round2(sum_tax),
    }
    dbg(box, f"  → BM税额={info['bm']}  PPN税额={info['ppn']}  PPH=不进表")
    dbg(box, f"  → 总税(逐行 H×M+(H+H×M)×O, 不含PPH)={info['total_ex_pph']}")
    return info


# ==================== 合同索引（性能核心）====================
def contract_name_key(filename):
    name = os.path.splitext(os.path.basename(filename))[0]
    name = re.sub(r"(?i)\s*_?iv\b", "", name)
    name = re.sub(r"(?i)\s*FORM\s*E", "", name)
    name = re.sub(r"[\(\)（）]", "", name)
    name = re.sub(r"^(975|总表)?\d*[_-]*", "", name, flags=re.I)
    name = name.strip("_- ")
    return norm(name)


def build_contract_index(folder, box=None):
    """
    ★ 只扫描 1 次目录，建 {关键词: 路径} 索引
    替代原来每一行都 glob 一次的高开销做法
    """
    ext_ok = (".xls", ".xlsx", ".xlsm")
    index = {}
    try:
        names = sorted(os.listdir(folder))
    except Exception as e:
        dbg(box, f"读取合同目录失败: {e}", force=True)
        return index

    valid = 0
    damaged = 0
    for name in names:
        if name.startswith("~$"):
            continue
        if os.path.splitext(name)[1].lower() not in ext_ok:
            continue
        p = os.path.join(folder, name)
        if not os.path.isfile(p):
            continue
        if not looks_like_xls_or_xlsx(p):
            dbg(box, f"  ⚠ 预检损坏（将报读取失败）: {name}")
            damaged += 1
        valid += 1
        base = os.path.splitext(name)[0]
        key = contract_name_key(base)
        raw = re.sub(r"[\s_\-()（）]", "", base).upper()
        if key:
            index.setdefault(key, p)
        if raw:
            index.setdefault(raw, p)

    dbg(box, f"合同索引建成：{valid} 个文件（其中疑似损坏 {damaged} 个），仅扫描 1 次", force=True)
    return index


def find_contract_fast(ivpl, index_dict, box=None):
    """内存查表匹配，O(1)；失败才走受限子串兜底"""
    if not ivpl:
        return None
    key = norm(ivpl)
    if key and key in index_dict:
        return index_dict[key]
    raw = re.sub(r"[\s_\-]", "", str(ivpl)).upper()
    if raw and raw in index_dict:
        return index_dict[raw]
    # 兜底：长度≥6 的子串互含
    for probe in (key, raw):
        if probe and len(probe) >= 6:
            for k, p in index_dict.items():
                if len(k) >= 6 and (k in probe or probe in k):
                    return p
    return None


def find_contract_file(folder, ivpl, box=None):
    """原 v2.6 匹配逻辑，保留作为索引为空时的兜底"""
    if not ivpl:
        return None
    key = norm(ivpl)
    if not key:
        return None
    candidates = [
        p for p in glob.glob(os.path.join(folder, "*"))
        if os.path.isfile(p) and os.path.splitext(p)[1].lower() in (".xls", ".xlsx", ".xlsm")
    ]
    exact = contain = None
    for p in candidates:
        k = contract_name_key(p)
        if k == key:
            exact = p
            break
        if contain is None and (key in k or k in key):
            contain = p
    hit = exact or contain
    if hit:
        dbg(box, f"  [匹配] 命中 -> {os.path.basename(hit)}")
        return hit
    raw = re.sub(r"[\s_\-]", "", str(ivpl)).upper()
    for p in candidates:
        base = re.sub(r"[\s_\-()（）]", "", os.path.splitext(os.path.basename(p))[0]).upper()
        if raw and raw in base:
            dbg(box, f"  [匹配] 兜底命中 -> {os.path.basename(p)}")
            return p
    dbg(box, f"  [匹配] 未找到（文件夹内 {len(candidates)} 个文件）")
    return None


# ==================== 主处理 ====================
def process(master_path, folder, box=None):
    dbg(box, "==== 开始 ====", force=True)
    dbg(box, f"总表: {os.path.basename(master_path)}")
    dbg(box, f"合同目录: {folder}")

    # ★ 1) 建索引（只扫 1 次）
    contract_index = build_contract_index(folder, box)
    use_index = len(contract_index) > 0

    # ★ 2) 总表必须普通模式（要写回保存），不能 read_only
    wb = load_workbook(master_path)
    ws = wb.active
    dbg(box, f"\n==== 总表 ====")
    dbg(box, f"Sheet: {ws.title}, 行={ws.max_row}, 列={ws.max_column}")

    headers_raw = [cell_text(ws.cell(HEADER_ROW, c).value) for c in range(1, ws.max_column + 1)]
    dbg(box, f"表头: {headers_raw}")

    ivpl_col   = col_by_names(ws, HEADER_ROW, [("IV&PL", "INVOICE NO", "发票号", "INVOICE", "合同号")])
    duty_col   = col_by_names(ws, HEADER_ROW, [("关税金额", "关税")])
    vat_col    = col_by_names(ws, HEADER_ROW, [("增值税金额", "增值税")])
    tax_col    = col_by_names(ws, HEADER_ROW, [("总税额", "总税")])
    remark_col = col_by_names(ws, HEADER_ROW, [("备注",)])
    dbg(box, f"列: IV&PL={ivpl_col} 关税={duty_col} 增值税={vat_col} 总税={tax_col} 备注={remark_col}", force=True)

    if not ivpl_col:
        dbg(box, "⚠️ 未找到 IV&PL 列！", force=True)

    processed = skipped = 0
    t0 = time.time()

    for r in range(DATA_START_ROW, ws.max_row + 1):
        ivpl = ws.cell(r, ivpl_col).value if ivpl_col else None
        if ivpl is None or str(ivpl).strip() == "":
            continue
        ivpl_s = cell_text(ivpl).strip()
        dbg(box, f"\n--- 第{r}行 发票号='{ivpl_s}' ---")

        # ★ 优先用索引查表；索引为空才回退原逻辑
        cp = find_contract_fast(ivpl_s, contract_index, box) if use_index \
            else find_contract_file(folder, ivpl_s, box)

        if not cp:
            if remark_col:
                ws.cell(r, remark_col).value = "未找到合同"
            skipped += 1
            continue

        dbg(box, f"  [匹配] 命中 -> {os.path.basename(cp)}")

        try:
            info = read_contract(cp, box)
        except Exception as e:
            dbg(box, f"  ⚠️ 读取异常: {e}", force=True)
            info = None

        if info is None:
            if remark_col:
                ws.cell(r, remark_col).value = "合同读取失败"
            skipped += 1
            continue

        # ① 关税金额 = Σ(H×M)
        if duty_col:
            ws.cell(r, duty_col).value = info["bm"]
        # ② 总税额 = Σ(H×M+(H+H×M)×O)，不含 PPH
        if tax_col:
            ws.cell(r, tax_col).value = info["total_ex_pph"]
        # ③ 增值税列：代码不填（总表公式 =ROUND((发票金额+关税)*0.11,2)），清残留旧值
        if vat_col:
            v = ws.cell(r, vat_col).value
            if v is not None and not str(v).lstrip().startswith("="):
                ws.cell(r, vat_col).value = None

        processed += 1
        dbg(box, f"  ✅ 关税(BM)={info['bm']}  总税(去PPH)={info['total_ex_pph']}")

    # ★ 输出带时间戳，避免 Excel 占用
    ts = datetime.now().strftime("%m%d_%H%M%S")
    stem = os.path.splitext(master_path)[0]
    out_path = f"{stem}_已填写_{ts}.xlsx"
    wb.save(out_path)
    try:
        wb.close()
    except Exception:
        pass

    dbg(box, f"\n完成！处理 {processed} 行，跳过 {skipped} 行，耗时 {time.time()-t0:.1f}s", force=True)
    dbg(box, f"输出: {out_path}", force=True)
    _flush_log(box)
    return out_path


# ==================== GUI ====================
def gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext

    root = tk.Tk()
    root.title("报关自动化工具 v2.8（分表修复 / 性能增强 / 标准四舍五入）")
    root.geometry("840x700")
    mv, fv = tk.StringVar(), tk.StringVar()

    tk.Label(root, text="① 选总表：").grid(row=0, column=0, sticky="e", padx=8, pady=12)
    tk.Entry(root, textvariable=mv, width=64).grid(row=0, column=1, padx=4)
    tk.Button(root, text="浏览…", command=lambda: mv.set(
        filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm *.xls")]) or mv.get())
    ).grid(row=0, column=2, padx=6)

    tk.Label(root, text="② 选合同文件夹：").grid(row=1, column=0, sticky="e", padx=8)
    tk.Entry(root, textvariable=fv, width=64).grid(row=1, column=1, padx=4)
    tk.Button(root, text="浏览…", command=lambda: fv.set(filedialog.askdirectory() or fv.get())
    ).grid(row=1, column=2, padx=6)

    box = scrolledtext.ScrolledText(root, height=34, font=("Consolas", 9))
    box.grid(row=3, column=0, columnspan=3, padx=12, pady=10, sticky="nsew")

    def run():
        if not mv.get() or not fv.get():
            return messagebox.showwarning("提示", "先选总表，再选合同文件夹")
        box.delete("1.0", "end")
        try:
            out = process(mv.get(), fv.get(), box)
            messagebox.showinfo("完成", f"已生成：\n{out}")
        except Exception:
            _flush_log(box)
            box.insert("end", traceback.format_exc())
            messagebox.showerror("报错", "请看日志")

    tk.Button(root, text="③ 开始处理", command=run, bg="#1f6feb", fg="white",
              font=("Microsoft YaHei", 10, "bold"), width=18).grid(row=2, column=1, pady=6)
    root.grid_rowconfigure(3, weight=1)
    root.grid_columnconfigure(1, weight=1)
    root.mainloop()


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        process(sys.argv[1], sys.argv[2], None)
    else:
        gui()
